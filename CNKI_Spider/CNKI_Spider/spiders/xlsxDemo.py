from openpyxl import Workbook,load_workbook

wb = Workbook()

# 激活 worksheet

ws = wb.active

# 数据可以直接分配到单元格中

ws['A1'] = '题名'
ws['B1'] = '作者'
ws['C1'] = '来源'
ws['D1'] = '发表时间'
ws['E1'] = '数据库'
ws['F1'] = '被引'
ws['G1'] = '下载'
ws['H1'] = '文章链接'
ws['I1'] = '摘要'


# 可以附加行，从第一列开始附加

# ws.append([1, 2, 3])

# Python 类型会被自动转换

import datetime
#
# ws['A3'] = datetime.datetime.now().strftime("%Y-%m-%d")

# 保存文件

wb.save("D:/Desktop/组织行为学.xlsx")